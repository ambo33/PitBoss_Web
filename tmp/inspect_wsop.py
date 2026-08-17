import json
import sys

import openpyxl

path = r"C:\Users\EricA\Downloads\2026 Road to the WSOP Standings.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

if len(sys.argv) >= 2 and sys.argv[1] == "dump":
    sheet_index = int(sys.argv[2])
    start = int(sys.argv[3])
    end = int(sys.argv[4])
    cols = int(sys.argv[5])
    ws = wb.worksheets[sheet_index]
    print(ws.title, ws.max_row, ws.max_column)
    for r in range(start, end + 1):
        vals = [ws.cell(r, c).value for c in range(1, cols + 1)]
        print(r, json.dumps(vals, default=str))
elif len(sys.argv) >= 2 and sys.argv[1] == "sheets":
    print(json.dumps([{"name": ws.title, "rows": ws.max_row, "cols": ws.max_column} for ws in wb.worksheets]))
elif len(sys.argv) >= 2 and sys.argv[1] == "parse-results":
    ws = wb.worksheets[0]
    players = []
    for index in range(36):
        col = 2 + (index * 3)
        name = ws.cell(1, col).value
        if name:
            players.append({"name": str(name).strip(), "col": col})
    league_fees = {}
    for player in players:
        league_fees[player["name"]] = ws.cell(2, player["col"]).value
    events = []
    for row in range(4, 13):
        event_name = ws.cell(row, 1).value
        event = {"name": str(event_name).strip(), "row": row, "results": []}
        for player in players:
            paid = ws.cell(row, player["col"]).value
            placed = ws.cell(row, player["col"] + 1).value
            points = ws.cell(row, player["col"] + 2).value
            dnf = str(placed).strip().lower() == "dnf"
            place = None if dnf else (int(placed) if isinstance(placed, (int, float)) else None)
            event["results"].append({
                "name": player["name"],
                "paid": paid,
                "place": place,
                "dnf": dnf,
                "points": int(points or 0),
            })
        events.append(event)
    points_lookup = []
    points_ws = wb.worksheets[2]
    for row in range(2, points_ws.max_row + 1):
        place_raw = points_ws.cell(row, 1).value
        points_raw = points_ws.cell(row, 2).value
        if place_raw is None:
            continue
        place = "DNF" if str(place_raw).strip().upper() == "DNF" else int(place_raw)
        points_lookup.append({"place": place, "points": int(points_raw or 0)})
    print(json.dumps({"players": players, "leagueFees": league_fees, "events": events, "pointsLookup": points_lookup}, default=str))
elif len(sys.argv) >= 2 and sys.argv[1] == "validate-results":
    ws = wb.worksheets[0]
    players = []
    for index in range(36):
        col = 2 + (index * 3)
        name = ws.cell(1, col).value
        if name:
            players.append({"name": str(name).strip(), "col": col})
    print("players", len(players))
    for row in range(4, 13):
        event_name = str(ws.cell(row, 1).value).strip()
        places = []
        dnfs = 0
        odd = []
        for player in players:
            placed = ws.cell(row, player["col"] + 1).value
            if str(placed).strip().lower() == "dnf":
                dnfs += 1
            elif isinstance(placed, (int, float)):
                places.append(int(placed))
            else:
                odd.append((player["name"], placed))
        dupes = sorted({place for place in places if places.count(place) > 1})
        missing = [place for place in range(1, len(places) + 1) if place not in places]
        print(event_name, "placed", len(places), "dnf", dnfs, "dupes", dupes or "-", "missing", missing or "-", "odd", odd or "-")
elif len(sys.argv) >= 2 and sys.argv[1] == "event-results":
    event_number = int(sys.argv[2])
    row = 3 + event_number
    ws = wb.worksheets[0]
    rows = []
    for index in range(36):
        col = 2 + (index * 3)
        name = str(ws.cell(1, col).value).strip()
        paid = ws.cell(row, col).value
        placed = ws.cell(row, col + 1).value
        points = ws.cell(row, col + 2).value
        rows.append({"name": name, "paid": paid, "placed": placed, "points": points})
    for row in sorted(rows, key=lambda item: (999 if str(item["placed"]).lower() == "dnf" else int(item["placed"]), item["name"])):
        print(json.dumps(row, default=str))
