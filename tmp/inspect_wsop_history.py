import json
import sys

import openpyxl


def cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def numeric(value):
    if isinstance(value, (int, float)):
        return float(value)
    try:
        raw = str(value).replace("$", "").replace(",", "").strip()
        return float(raw) if raw else None
    except (TypeError, ValueError):
        return None


def common_number(values, default=0):
    counts = {}
    for value in values:
        number = numeric(value)
        if number is None:
            continue
        key = round(number, 2)
        counts[key] = counts.get(key, 0) + 1
    if not counts:
        return default
    return sorted(counts.items(), key=lambda item: (-item[1], item[0]))[0][0]


def workbook_payload(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    return {
        "sheets": [
            {"index": index, "name": ws.title, "rows": ws.max_row, "cols": ws.max_column}
            for index, ws in enumerate(wb.worksheets)
        ]
    }


def dump(path, sheet_index, start, end, cols):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[sheet_index]
    rows = []
    for row in range(start, end + 1):
        rows.append({
            "row": row,
            "values": [ws.cell(row, col).value for col in range(1, cols + 1)],
        })
    return {"sheet": ws.title, "rows": rows}


def find_players(ws):
    players = []
    # Historical sheets use a repeating three-column player block:
    # Paid / Place / Points, with the player name in the first row of the block.
    for col in range(2, ws.max_column + 1, 3):
        name = cell_text(ws.cell(1, col).value)
        if name and name.lower() not in {"expenses", "total", "totals"}:
            players.append({"name": name, "col": col})
    return players


def parse_results(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    players = find_players(ws)
    league_fees = {player["name"]: ws.cell(2, player["col"]).value for player in players}
    league_fee = common_number(league_fees.values())

    events = []
    for row in range(4, ws.max_row + 1):
        event_name = cell_text(ws.cell(row, 1).value)
        if not event_name:
            continue
        # Stop when we leave the compact event-result block.
        if event_name.lower().startswith(("show bonus", "top ", "total", "rank", "final", "place", "name")):
            break
        results = []
        saw_result = False
        for player in players:
            paid = ws.cell(row, player["col"]).value
            placed = ws.cell(row, player["col"] + 1).value
            points = ws.cell(row, player["col"] + 2).value
            if paid is not None or placed is not None or points is not None:
                saw_result = True
            dnf = cell_text(placed).lower() == "dnf"
            place = None
            if not dnf and isinstance(placed, (int, float)):
                place = int(placed)
            results.append({
                "name": player["name"],
                "paid": paid,
                "place": place,
                "dnf": dnf,
                "points": int(points or 0),
            })
        if saw_result:
            events.append({"name": event_name, "row": row, "results": results})

    event_fee = common_number(
        result["paid"]
        for event in events
        for result in event["results"]
        if cell_text(result["paid"]).lower() != "dnf"
    )

    show_bonus_row = None
    for row in range(1, ws.max_row + 1):
        label = cell_text(ws.cell(row, 1).value).lower()
        if label.startswith("show bonus"):
            show_bonus_row = row
            break
    show_bonus_candidates = []
    if show_bonus_row:
        for player in players:
            total_bonus = numeric(ws.cell(show_bonus_row, player["col"]).value)
            played = 0
            for event in events:
                player_result = next((result for result in event["results"] if result["name"] == player["name"]), None)
                if player_result and not player_result["dnf"] and player_result["place"] is not None:
                    played += 1
            if total_bonus is not None and played:
                show_bonus_candidates.append(total_bonus / played)
    showup_bonus = int(round(common_number(show_bonus_candidates, 0)))

    points_lookup = []
    for points_ws in wb.worksheets[1:]:
        for row in range(1, points_ws.max_row + 1):
            place_raw = points_ws.cell(row, 1).value
            points_raw = points_ws.cell(row, 2).value
            if place_raw is None or points_raw is None:
                continue
            place_text = cell_text(place_raw)
            if place_text.lower() in {"place", "rank"}:
                continue
            try:
                place = "DNF" if place_text.upper() == "DNF" else int(float(place_raw))
                points = int(float(points_raw))
            except (TypeError, ValueError):
                continue
            points_lookup.append({"place": place, "points": points})
        if points_lookup:
            break

    if points_lookup and not any(rule["place"] == "DNF" for rule in points_lookup):
        points_lookup.insert(0, {"place": "DNF", "points": 0})

    return {
        "players": players,
        "leagueFees": league_fees,
        "settings": {
            "expectedPlayerCount": len(players),
            "leagueFee": league_fee,
            "perEventFee": event_fee,
            "showupBonusPoints": showup_bonus,
            "bestFinishCount": 7,
        },
        "events": events,
        "pointsLookup": points_lookup,
    }


if __name__ == "__main__":
    command = sys.argv[1]
    path = sys.argv[2]
    if command == "sheets":
        print(json.dumps(workbook_payload(path), default=str))
    elif command == "dump":
        print(json.dumps(dump(path, int(sys.argv[3]), int(sys.argv[4]), int(sys.argv[5]), int(sys.argv[6])), default=str))
    elif command == "parse-results":
        print(json.dumps(parse_results(path), default=str))
    else:
        raise SystemExit(f"Unknown command: {command}")
